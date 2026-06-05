from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


path = Path("/Users/wangsiyan/Courses/面计离散/proj/outputs/v7_structured_workbook/experiment_v7_structured_results.xlsx")
wb = load_workbook(path)

if "Charts" in wb.sheetnames:
    del wb["Charts"]

ws_a = wb["Experiment_A_Small"]
ws_b = wb["Experiment_B_Large"]
ws = wb.create_sheet("Charts")

title_fill = PatternFill("solid", fgColor="DDEBF0")
header_fill = PatternFill("solid", fgColor="1F4E5F")
header_font = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="D8E2E7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_title(cell, text):
    ws[cell] = text
    ws[cell].font = Font(bold=True, size=16, color="173A46")
    ws[cell].fill = title_fill


def write_table(start_row, start_col, headers, rows):
    for offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for r_offset, row in enumerate(rows, start=1):
        for c_offset, value in enumerate(row):
            cell = ws.cell(start_row + r_offset, start_col + c_offset, value)
            cell.border = border
            if isinstance(value, (int, float)):
                cell.number_format = "0.00"


labels_a = ["N=8 No Trap", "N=8 Trap", "N=10 No Trap", "N=10 Trap"]
labels_b = ["N=50 No Trap", "N=50 Trap", "N=100 No Trap", "N=100 Trap", "N=200 No Trap", "N=200 Trap"]

write_title("A1", "Experiment v7 Charts")

exp_a_ratio = []
for i, label in enumerate(labels_a, start=4):
    exp_a_ratio.append([label, ws_a.cell(i, 10).value, ws_a.cell(i, 17).value, ws_a.cell(i, 23).value])
write_table(3, 1, ["Condition", "Lap6/Base (%)", "Lap8/Base (%)", "SVD8/Base (%)"], exp_a_ratio)

exp_a_time = []
for i, label in enumerate(labels_a, start=4):
    exp_a_time.append([label, ws_a.cell(i, 5).value, ws_a.cell(i, 12).value, ws_a.cell(i, 19).value, ws_a.cell(i, 25).value])
write_table(3, 6, ["Condition", "Base Time (ms)", "Lap6 Time (ms)", "Lap8 Time (ms)", "SVD8 Time (ms)"], exp_a_time)

exp_b_ratio = []
for i, label in enumerate(labels_b, start=4):
    exp_b_ratio.append([label, ws_b.cell(i, 8).value, ws_b.cell(i, 15).value, ws_b.cell(i, 21).value])
write_table(26, 1, ["Condition", "Lap6/Theory6 (%)", "Lap8/Theory8 (%)", "SVD8/Theory8 (%)"], exp_b_ratio)

exp_b_time = []
for i, label in enumerate(labels_b, start=4):
    exp_b_time.append([label, ws_b.cell(i, 10).value, ws_b.cell(i, 17).value, ws_b.cell(i, 23).value])
write_table(26, 6, ["Condition", "Lap6 Time (ms)", "Lap8 Time (ms)", "SVD8 Time (ms)"], exp_b_time)


def add_bar_chart(title, anchor, min_col, max_col, min_row, max_row, y_axis_title, log_scale=False):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = "Condition"
    chart.width = 21
    chart.height = 11
    chart.legend.position = "b"
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True
    chart.dataLabels.showLegendKey = False
    chart.dataLabels.showCatName = False
    chart.dataLabels.showSerName = False
    chart.y_axis.majorGridlines = ChartLines()
    if log_scale:
        chart.y_axis.scaling.logBase = 10
        chart.y_axis.scaling.min = 0.1
        chart.y_axis.scaling.max = 10000
        chart.x_axis.crosses = "min"
    data = Reference(ws, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=min_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


add_bar_chart("Experiment A: Profit Ratio vs Base", "A10", 1, 4, 3, 7, "Ratio (%)")
add_bar_chart("Experiment A: Runtime Comparison", "J10", 6, 10, 3, 7, "Time (ms, log scale)", log_scale=True)
add_bar_chart("Experiment B: Profit Ratio vs Theory", "A34", 1, 4, 26, 32, "Ratio (%)")
add_bar_chart("Experiment B: Pruned Runtime Comparison", "J34", 6, 9, 26, 32, "Time (ms)")

for col in range(1, 11):
    ws.column_dimensions[chr(64 + col)].width = 18

wb.save(path)
print(path)
